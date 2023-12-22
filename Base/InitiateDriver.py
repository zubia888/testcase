from Library import ConfigFileReader
from selenium import webdriver
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.webdriver.chrome.service import Service
import openpyxl
import time
import datetime
import os


def start_browser():
    global driver
    if ((ConfigFileReader.readConfigData('Details', 'Browser')) == 'Chrome'):
        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(service=Service(), options=options)
    elif ((ConfigFileReader.readConfigData('Details', 'Browser')) == 'Firefox'):
        path_binary = "C:\\Program Files (x86)\\Mozilla Firefox\\firefox.exe"
        firefox_binary = FirefoxBinary(path_binary)
        options = webdriver.FirefoxOptions()
        options.binary = firefox_binary
        driver = webdriver.Firefox(options=options)
    else:
        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(service=Service(), options=options)
    driver.get(ConfigFileReader.readConfigData('Details', 'Application_URL'))
    # driver.maximize_window()
    return driver


def date_time():
    now = datetime.datetime.now()
    date_time = now.strftime("%m-%d-%Y %Hh %Mm %Ss")
    return date_time


def screen_shot():
    new_folder_name = "test_SS"
    os.makedirs(new_folder_name, exist_ok=True)
    screenshot_path = os.path.join(new_folder_name, f"before registration - {date_time()}.png")
    driver.get_screenshot_as_file(screenshot_path)
    time.sleep(5)


def create_sheet(data):
    file_path = 'registration_data_file.xlsx'
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        last_row = sheet.max_row + 1
        for i in range(len(data)):
            sheet[f'{chr(65 + i)}{last_row}'] = data[i]
        workbook.save(file_path)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        header = ['Test Run', 'User', 'Email', 'Password', 'Confirm Password', 'DOB', 'Phone No.', 'Address',
                  'Gender', 'Country', 'State', 'City', 'Zip Code', 'Time']
        for col_num, label in enumerate(header, start=1):
            sheet.cell(row=1, column=col_num, value=label)
        workbook.save(file_path)

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        last_row = sheet.max_row + 1
        for i in range(len(data)):
            sheet[f'{chr(65 + i)}{last_row}'] = data[i]
        workbook.save(file_path)


def save_data(data):
    # Create Text File
    with open("registration_data_file.txt", "a+") as f:
        # f.writelines(f"Test Run #. {run_count + 1}\n")
        f.writelines([str(item) + "\n" for item in data])
        # f.writelines(f"Time: {date_time()}")
        f.writelines("\n \n")

    # Count Increment
    # increment_and_save_count()


def close_browser():
    driver.close()
